"""
excel_output.py — Build the Excel deliverable from analysis results.

Produces a workbook with:
  Sheet 1: 'Analysis' — the main table the user requested:
             1. Your Sloka
             2. Splitting of Words
             3. English Meanings of Words
             4. Line-by-Line Meaning

  Sheet 2: 'Summary' — a single-sheet overview (counts, coverage stats)
  Sheet 3: 'About' — explains what the SLM did, so users/reviewers understand
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _format_word_splits(word_analyses: list) -> str:
    """Convert word analyses into a readable 'splitting' string.

    Example output:
      "वातपित्तकफा: vāta + pitta; दोषाः = doṣa; समासेन = samāsena"
    """
    parts = []
    for word, subs in word_analyses:
        if not subs:
            parts.append(f"{word}: [no match]")
        elif len(subs) == 1 and subs[0][0] == word:
            iast = subs[0][1].get('iast', '')
            parts.append(f"{word} = {iast}" if iast else word)
        else:
            sub_iasts = [s[1].get('iast', s[0]) for s in subs]
            parts.append(f"{word}: {' + '.join(sub_iasts)}")
    return " | ".join(parts)


def _format_word_meanings(word_analyses: list) -> str:
    """Word-by-word English meanings as a single string.

    Example output:
      "वातपित्तकफा: Vāta — ...; Pitta — ... | दोषाः = Doṣa — regulatory factor"
    """
    parts = []
    for word, subs in word_analyses:
        if not subs:
            parts.append(f"{word}: [no dictionary match]")
        elif len(subs) == 1 and subs[0][0] == word:
            parts.append(f"{word} = {subs[0][1]['english']}")
        else:
            sub_meanings = [f"{s[0]}: {s[1]['english']}" for s in subs]
            parts.append(f"{word} → " + "; ".join(sub_meanings))
    return "\n".join(parts)


def _style_header_row(ws, row_num: int, ncols: int) -> None:
    """Style the header row: bold, centered, dark fill, white text."""
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border


def _style_data_cells(ws, start_row: int, end_row: int, ncols: int) -> None:
    """Apply wrap-text, borders, and alignment to data cells."""
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align
            cell.border = border


def build_excel(
    results: list[dict],
    output_path: str | None = None,
) -> bytes:
    """
    Build the Excel output from a list of analysis results.

    Each `result` is the dict returned by CharakaAnalyzer.analyze_sloka():
      {
        'input': str,
        'words': [(word, [(sub, info), ...]), ...],
        'english_gloss': str,
        'similar': [{'chapter': int, 'sloka_num': int, 'text': str, 'similarity': float}, ...]
      }

    Returns the Excel file bytes (also writes to output_path if provided).
    """
    wb = Workbook()

    # ---------------- Sheet 1: Analysis ----------------
    ws = wb.active
    ws.title = 'Analysis'

    headers = [
        '#',
        'Your Sloka',
        'Splitting of Words',
        'English Meanings of Words',
        'Line-by-Line Meaning',
    ]
    ws.append(headers)

    # Coverage tracking
    total_words = 0
    matched_words = 0
    compound_count = 0

    for idx, r in enumerate(results, 1):
        split_str = _format_word_splits(r['words'])
        meanings_str = _format_word_meanings(r['words'])
        line_meaning = r['english_gloss']

        ws.append([idx, r['input'], split_str, meanings_str, line_meaning])

        # Track coverage
        for word, subs in r['words']:
            total_words += 1
            if subs:
                matched_words += 1
                if len(subs) > 1 or (subs and subs[0][0] != word):
                    compound_count += 1

    n_data_rows = len(results)
    _style_header_row(ws, 1, len(headers))
    _style_data_cells(ws, 2, 1 + n_data_rows, len(headers))

    # Column widths
    ws.column_dimensions['A'].width = 5    # #
    ws.column_dimensions['B'].width = 45   # Your Sloka
    ws.column_dimensions['C'].width = 40   # Splitting
    ws.column_dimensions['D'].width = 60   # Meanings
    ws.column_dimensions['E'].width = 50   # Line meaning

    # Row heights: enough to show wrapped content
    ws.row_dimensions[1].height = 30
    for r in range(2, 2 + n_data_rows):
        ws.row_dimensions[r].height = 90

    # Freeze header
    ws.freeze_panes = 'A2'

    # ---------------- Sheet 2: Summary ----------------
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 30

    summary_font = Font(bold=True, size=14, color='2C3E50', name='Calibri')
    ws2['A1'] = 'Charaka Analysis — Summary'
    ws2['A1'].font = summary_font
    ws2.merge_cells('A1:B1')

    coverage = (matched_words / total_words * 100) if total_words else 0
    summary_rows = [
        ('', ''),
        ('Generated on', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Total ślokas analyzed', len(results)),
        ('Total words', total_words),
        ('Words matched in dictionary', matched_words),
        ('Dictionary coverage', f'{coverage:.1f}%'),
        ('Compound words decomposed', compound_count),
        ('', ''),
        ('Model used', 'Custom Sanskrit SLM (1.3M-parameter transformer)'),
        ('Corpus for retrieval', '1,968 ślokas from Charaka Saṃhitā Sūtrasthāna'),
        ('Tokenizer', 'SentencePiece BPE (4,000 tokens, Sanskrit-trained)'),
        ('Dictionary', 'WHO Ayurveda Terminology + curated canonical-meaning supplement'),
    ]
    for i, (key, val) in enumerate(summary_rows, start=2):
        ws2[f'A{i}'] = key
        ws2[f'B{i}'] = str(val)
        if key:
            ws2[f'A{i}'].font = Font(bold=True, name='Calibri')

    # ---------------- Sheet 3: About ----------------
    ws3 = wb.create_sheet('About')
    ws3.column_dimensions['A'].width = 100
    ws3['A1'] = 'About this tool'
    ws3['A1'].font = Font(bold=True, size=14, color='2C3E50', name='Calibri')

    about_text = [
        '',
        'This Excel was generated by the Charaka Sanskrit SLM App.',
        '',
        'How the analysis works (4 stages, all offline):',
        '',
        '1. INPUT: Your śloka (from text input, image OCR, PDF, DOCX, or Excel)',
        '',
        '2. TOKENIZATION: A SentencePiece BPE tokenizer (vocab 4,000) that was trained',
        '   specifically on Charaka Saṃhitā Sūtrasthāna splits the śloka into sub-word',
        '   pieces that respect Sanskrit sandhi and compound structure.',
        '',
        '3. DICTIONARY LOOKUP + COMPOUND DECOMPOSITION:',
        '   For each word, the app tries a direct lookup against:',
        '      a) a curated canonical-meaning supplement (~150 entries)',
        '      b) the WHO Ayurveda Terminology (~1,500 clinical terms)',
        '   For unmatched compounds, a dynamic-programming best-cover algorithm',
        '   finds the optimal decomposition into known sub-terms.',
        '',
        '4. SEMANTIC RETRIEVAL: A 1.3M-parameter transformer encoder (trained via',
        '   masked language modeling on the Charaka corpus) produces a 128-dim',
        '   sentence embedding for each input śloka. This embedding is compared',
        '   against 1,968 precomputed corpus embeddings by cosine similarity.',
        '',
        'Nothing in the pipeline uses an external API. All inference runs locally',
        'on CPU. Output quality depends on OCR accuracy (for image/PDF inputs),',
        'which you can review and correct before analysis.',
        '',
        'Developed by Dr. Prasanna Kulkarni (Atharva AyurTech Healthcare).',
    ]
    for i, line in enumerate(about_text, start=2):
        ws3[f'A{i}'] = line
        ws3[f'A{i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ---------------- Save ----------------
    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(xlsx_bytes)
    return xlsx_bytes
